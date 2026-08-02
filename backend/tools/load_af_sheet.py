"""
Load the latest operational workbook (AF Sheet.xlsx) into MongoDB Atlas and stamp
the Indian Financial Year on every record.

Why
---
`AF Sheet.xlsx` is the workbook the team actually keys into daily, and it runs
ahead of both the Google Sheet and Atlas (143 quotes for 2026 vs 117 loaded).
This pulls it in and, crucially, tags each row with its FY so the app can hide
or show a year's content.

Financial year (India): 1 April -> 31 March.  FY label = "2026-27".

Hygiene rules carried over from the web CRM (learned the hard way):
  * header is NOT on row 1 (title + colour-legend rows sit above it) -> detected
  * phones arrive as floats ("9158668737.0") -> digits only, last 10
  * dates are wildly inconsistent ("04//01/26", "24/o1/2026") -> best effort
  * a value > 1e10 is a timestamp that leaked into a money column -> zeroed
  * blank natural key = junk row -> skipped

Usage:
    python backend/tools/load_af_sheet.py --dry-run
    python backend/tools/load_af_sheet.py
"""
import asyncio
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND))

env_path = BACKEND / ".env"
if env_path.exists():
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

import openpyxl  # noqa: E402
from motor.motor_asyncio import AsyncIOMotorClient  # noqa: E402

XLSX = Path(r"C:\Users\jagad\Downloads\AF Sheet.xlsx")

HDR_TOKENS = ["quot", "phone", "mobile", "s.no", "sl", "client", "cust", "name",
              "date", "status", "sale", "refer"]


# ── financial year ─────────────────────────────────────────────────────────
def fy_of(iso: str) -> str:
    """'2026-05-14' -> '2026-27'.  Indian FY starts 1 April."""
    if not iso or len(iso) < 7:
        return ""
    try:
        y, m = int(iso[:4]), int(iso[5:7])
    except ValueError:
        return ""
    start = y if m >= 4 else y - 1
    return f"{start}-{str(start + 1)[-2:]}"


def s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v).strip()


def phone(v):
    d = re.sub(r"\D", "", s(v))
    return d[-10:] if len(d) > 10 else d


def money(v):
    try:
        n = float(re.sub(r"[^\d.\-]", "", s(v)) or 0)
    except ValueError:
        return 0.0
    return 0.0 if abs(n) > 1e10 else n


def date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    t = s(v)
    if not t:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}", t):
        return t[:10]
    t = t.replace("o", "0").replace("O", "0")          # '24/o1/2026'
    m = re.search(r"(\d{1,2})\s*[/\-. ]+\s*(\d{1,2})\s*[/\-. ]+\s*(\d{2,4})", t)
    if m:
        d_, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d_).date().isoformat()
        except ValueError:
            return ""
    return ""


def header_row(rows):
    bi, bs = 0, -1
    for i, r in enumerate(rows[:8]):
        cells = [str(c).lower() for c in r if c is not None]
        sc = sum(any(t in c for t in HDR_TOKENS) for c in cells)
        if sc > bs:
            bi, bs = i, sc
    return bi


def read_sheet(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    h = header_row(rows)
    hdr = [s(c) for c in rows[h]]
    data = [r for r in rows[h + 1:] if any(c is not None and s(c) for c in r)]
    return hdr, data


def col(hdr, *keys):
    for j, c in enumerate(hdr):
        cl = c.lower()
        if any(k in cl for k in keys):
            return j
    return -1


def get(row, j):
    return row[j] if 0 <= j < len(row) else None


def now_iso():
    return datetime.now(timezone.utc).isoformat()


async def main(dry: bool):
    if not XLSX.exists():
        print(f"workbook not found: {XLSX}")
        return
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    db = AsyncIOMotorClient(os.environ.get("MONGO_URL"), serverSelectionTimeoutMS=20000)[
        os.environ.get("DB_NAME", "madio_crm")
    ]
    await db.client.admin.command("ping")
    print("DRY RUN — nothing written\n" if dry else "WRITING to Atlas\n")

    totals = {}

    # ── quotes: all three year tabs ────────────────────────────────────────
    quotes = []
    for tab in ("2026 MF Quotes", "2025 MF Quotes", "2024 MF Quotes"):
        if tab not in wb.sheetnames:
            continue
        hdr, data = read_sheet(wb, tab)
        jq = col(hdr, "quot. no", "quot no")
        jd = col(hdr, "quot. date", "client visit")
        jn = col(hdr, "cust name", "client name")
        jr = col(hdr, "reference")
        jp = col(hdr, "phone", "mobile")
        jreq = col(hdr, "requirement")
        jby = col(hdr, "attend")
        jst = col(hdr, "status")
        for r in data:
            qno = s(get(r, jq))
            if not qno:
                continue
            dt = date(get(r, jd))
            quotes.append({
                "quote_no": qno,
                "date": dt,
                "fy": fy_of(dt),
                "customer": s(get(r, jn)),
                "reference": s(get(r, jr)),
                "phone": phone(get(r, jp)),
                "division": "Furniture",
                "by_user": s(get(r, jby)),
                "stage": (s(get(r, jst)) or "Quoted")[:40],
                "remarks": s(get(r, jreq)),
            })
    totals["quotes"] = quotes

    # ── sales ──────────────────────────────────────────────────────────────
    hdr, data = read_sheet(wb, "MF Sale")
    js = col(hdr, "sale no")
    jd = col(hdr, "adv / cof", "adv/cof", "date")
    jn = col(hdr, "cust name")
    jq = col(hdr, "quote no")
    jp = col(hdr, "phone")
    jby = col(hdr, "attend")
    sales = []
    for r in data:
        sno = s(get(r, js))
        if not sno:
            continue
        dt = date(get(r, jd))
        sales.append({
            "sale_no": sno, "date": dt, "fy": fy_of(dt),
            "customer": s(get(r, jn)), "division": "Furniture",
            "quote_ref": s(get(r, jq)), "by_user": s(get(r, jby)),
            "phone": phone(get(r, jp)),
        })
    totals["sales"] = sales

    # ── visitors ───────────────────────────────────────────────────────────
    hdr, data = read_sheet(wb, "Visitors")
    jd = col(hdr, "date")
    jn = col(hdr, "cust name")
    jr = col(hdr, "reference")
    jp = col(hdr, "phone")
    jreq = col(hdr, "requirement")
    jby = col(hdr, "attend")
    visitors = []
    for r in data:
        nm = s(get(r, jn))
        if not nm:
            continue
        dt = date(get(r, jd))
        visitors.append({
            "date": dt, "fy": fy_of(dt), "name": nm,
            "reference": s(get(r, jr)), "phone": phone(get(r, jp)),
            "requirement": s(get(r, jreq)), "attend_person": s(get(r, jby)),
            "stage": "New Lead",
        })
    totals["visitors"] = visitors

    # ── navaki -> leads ────────────────────────────────────────────────────
    hdr, data = read_sheet(wb, "Navaki")
    jd = col(hdr, "date")
    jn = col(hdr, "client name", "names")
    jp = col(hdr, "mobile", "phone")
    jr = col(hdr, "refer")
    leads = []
    for r in data:
        nm = s(get(r, jn))
        if not nm:
            continue
        dt = date(get(r, jd))
        leads.append({
            "date": dt, "fy": fy_of(dt), "name": nm,
            "phone": phone(get(r, jp)), "source": "Navaki",
            "stage": "New", "remarks": s(get(r, jr)),
        })
    totals["leads"] = leads

    KEYS = {"quotes": ("quote_no",), "sales": ("sale_no",),
            "visitors": ("name", "phone", "date"), "leads": ("name", "phone", "date")}

    for coll, docs in totals.items():
        before = await db[coll].count_documents({})
        ins = upd = 0
        if not dry:
            for d in docs:
                q = {k: d.get(k, "") for k in KEYS[coll]}
                ex = await db[coll].find_one(q, {"_id": 1})
                if ex:
                    await db[coll].update_one({"_id": ex["_id"]}, {"$set": d})
                    upd += 1
                else:
                    await db[coll].insert_one({"id": str(uuid.uuid4()), **d, "created_at": now_iso()})
                    ins += 1
        after = await db[coll].count_documents({})
        print(f"  {coll:<10} parsed={len(docs):<5} inserted={ins:<5} updated={upd:<5} total {before} -> {after}")

    # FY distribution across the whole database
    if not dry:
        print("\nfinancial-year spread:")
        for coll in totals:
            pipe = [{"$group": {"_id": "$fy", "n": {"$sum": 1}}}, {"$sort": {"_id": -1}}]
            got = {(d["_id"] or "(none)"): d["n"] async for d in db[coll].aggregate(pipe)}
            print(f"  {coll:<10} {got}")
    wb.close()


if __name__ == "__main__":
    asyncio.run(main("--dry-run" in sys.argv))
